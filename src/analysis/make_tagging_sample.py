"""
make_tagging_sample.py - 정확도 검증용 태깅 샘플 추출
======================================================
naver_samsung_news_labeled.csv에서 100건을 샘플링해서
태깅하기 쉬운 엑셀 파일로 만들어줍니다.

실행: python make_tagging_sample.py
출력: tagging_sample.xlsx
"""

import pandas as pd
import numpy as np

INPUT_CSV  = "naver_samsung_news_labeled.csv"
OUTPUT_XLS = "tagging_sample.xlsx"
SAMPLE_N   = 100  # 태깅할 샘플 수

# ============================================================
# 데이터 로드
# ============================================================
df = pd.read_csv(INPUT_CSV, encoding="utf-8-sig")
print(f"전체 데이터: {len(df)}건")

# ============================================================
# 균형 샘플링 (긍정/부정/중립 균등하게)
# 긍정 34건, 부정 33건, 중립 33건
# ============================================================
pos = df[df["sentiment"] == "positive"].sample(n=34, random_state=42)
neg = df[df["sentiment"] == "negative"].sample(n=33, random_state=42)
neu = df[df["sentiment"] == "neutral" ].sample(n=33, random_state=42)

sample = pd.concat([pos, neg, neu]).sample(frac=1, random_state=42).reset_index(drop=True)
sample.index += 1  # 1부터 시작

# ============================================================
# 태깅용 컬럼 구성
# ============================================================
output = pd.DataFrame({
    "번호":           sample.index,
    "제목":           sample["title"],
    "본문앞200자":    sample["content"].str[:200],
    "모델예측":       sample["sentiment"],
    "확신도":         sample["sentiment_score"],
    "내태깅":         "",   # ← 여기에 직접 입력 (positive/negative/neutral)
    "일치여부":       "",   # ← 자동 계산됨
})

# ============================================================
# 엑셀 저장 (서식 포함)
# ============================================================
with pd.ExcelWriter(OUTPUT_XLS, engine="openpyxl") as writer:
    output.to_excel(writer, index=False, sheet_name="태깅")

    ws = writer.sheets["태깅"]

    # 열 너비 조정
    ws.column_dimensions["A"].width = 6   # 번호
    ws.column_dimensions["B"].width = 45  # 제목
    ws.column_dimensions["C"].width = 60  # 본문앞200자
    ws.column_dimensions["D"].width = 12  # 모델예측
    ws.column_dimensions["E"].width = 8   # 확신도
    ws.column_dimensions["F"].width = 15  # 내태깅
    ws.column_dimensions["G"].width = 10  # 일치여부

    # 헤더 색상
    from openpyxl.styles import PatternFill, Font, Alignment
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 행 높이 + 텍스트 줄바꿈
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 60
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 모델예측 색상 (긍정=파랑, 부정=빨강, 중립=회색)
    from openpyxl.styles import Font as OFont
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if cell.value == "positive":
                cell.font = OFont(color="1F4E79", bold=True)
            elif cell.value == "negative":
                cell.font = OFont(color="C00000", bold=True)
            else:
                cell.font = OFont(color="595959")

print(f"\n저장 완료: {OUTPUT_XLS}")
print(f"샘플 수: {len(output)}건")
print(f"  긍정: {len(pos)}건")
print(f"  부정: {len(neg)}건")
print(f"  중립: {len(neu)}건")
print("\n=== 사용 방법 ===")
print("1. tagging_sample.xlsx 열기")
print("2. '내태깅' 열에 각 기사를 읽고 직접 입력:")
print("   긍정 → positive")
print("   부정 → negative")
print("   중립 → neutral")
print("3. 100건 태깅 완료 후 evaluate.py 실행")